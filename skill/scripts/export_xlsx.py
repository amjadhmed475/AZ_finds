#!/usr/bin/env python3
"""Export research JSON to a multi-sheet Excel workbook.

Accepts either a dashboard JSON (with `products[].profitability`) or the
export payload JSON written by the MCP `export_product_report` tool
(`{products, profitability, suppliers}`).

Usage:
    python export_xlsx.py <research.json> [out.xlsx]

Requires: openpyxl  (pip install openpyxl)
"""
import json
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Install openpyxl first:  pip install openpyxl", file=sys.stderr)
    raise SystemExit(1)


def load(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def prof_index(data):
    idx = {}
    for p in data.get("profitability", []):
        idx[p.get("product_candidate_id")] = p
    return idx


def main() -> int:
    if len(sys.argv) < 2:
        print("usage: export_xlsx.py <research.json> [out.xlsx]", file=sys.stderr)
        return 2
    data = load(sys.argv[1])
    out = sys.argv[2] if len(sys.argv) > 2 else "amazon-research.xlsx"

    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="1F2937")
    head_font = Font(color="FFFFFF", bold=True)

    # Sheet 1: ranking
    ws = wb.active
    ws.title = "Products"
    cols = ["#", "Product", "Category", "Sale $", "Monthly Sales", "Landed $",
            "Net/Unit $", "ROI %", "Margin %", "Sellers", "Rating", "Trend",
            "Gating", "Risk", "Score", "Action"]
    ws.append(cols)
    for c in ws[1]:
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center")

    products = data.get("products", [])
    pidx = prof_index(data)
    for i, p in enumerate(products, 1):
        prof = p.get("profitability") or pidx.get(p.get("id"), {})
        ws.append([
            i, p.get("name", ""), p.get("category", ""),
            p.get("estimatedSalePrice", p.get("estimated_sale_price", "")),
            p.get("estimatedMonthlySales", p.get("estimated_monthly_sales", "")),
            prof.get("landedCost", prof.get("landed_cost", "")),
            prof.get("netProfit", prof.get("net_profit", "")),
            prof.get("roi", prof.get("roi_percent", "")),
            prof.get("margin", prof.get("net_margin_percent", "")),
            p.get("seller_count", p.get("sellerCount", "")),
            p.get("reviewRating", p.get("review_rating_average", "")),
            p.get("trend_status", p.get("trendStatus", "")),
            p.get("gating_status", p.get("gatingStatus", "")),
            p.get("risk_level", p.get("riskLevel", "")),
            p.get("opportunity_score", p.get("opportunityScore", "")),
            p.get("recommended_action", p.get("recommendedAction", "")),
        ])
    for col in ws.columns:
        width = max(10, min(40, max(len(str(c.value or "")) for c in col) + 2))
        ws.column_dimensions[col[0].column_letter].width = width

    # Sheet 2: suppliers
    sup = wb.create_sheet("Suppliers")
    scols = ["Product", "Supplier", "Site", "Region", "Unit $ min", "Unit $ max",
             "MOQ", "Landed $", "Rating", "Lead time", "Match", "Risk", "URL"]
    sup.append(scols)
    for c in sup[1]:
        c.fill = head_fill
        c.font = head_font
    suppliers = data.get("suppliers", [])
    if not suppliers:
        for p in products:
            for s in p.get("suppliers", []):
                suppliers.append(s)
    for s in suppliers:
        sup.append([
            s.get("product_candidate_id", ""), s.get("supplier_name", ""),
            s.get("supplier_site", ""), s.get("region", ""),
            s.get("unit_cost_min", ""), s.get("unit_cost_max", ""),
            s.get("moq", ""), s.get("estimated_landed_cost", ""),
            s.get("supplier_rating", ""), s.get("lead_time", ""),
            s.get("match_quality_score", ""), s.get("supplier_risk", ""),
            s.get("direct_product_url", ""),
        ])

    # Sheet 3: notes
    notes = wb.create_sheet("Read Me")
    notes["A1"] = "All figures are ESTIMATES. Verify every product in Amazon Seller Central."
    notes["A2"] = "Gating is an estimate, never a guarantee. Treat as 'manual verification required'."
    notes["A1"].font = Font(bold=True, color="B91C1C")

    wb.save(out)
    print(f"Wrote {out} with {len(products)} products and {len(suppliers)} supplier rows.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
